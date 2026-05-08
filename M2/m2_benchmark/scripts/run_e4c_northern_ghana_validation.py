from __future__ import annotations

import argparse
import hashlib
import json
import sys
import time
from dataclasses import asdict
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[3]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from hydrosheaf.api import fit_network_pipeline
from hydrosheaf.config import default_config
from hydrosheaf.graph.types import Edge
from hydrosheaf.inference.network_fit import infer_edges


BENCHMARK_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / "data" / "NorthenGhana"
WORKBOOK_PATH = SOURCE_DIR / "NorthernGhana.xlsx"
SI_PATH = SOURCE_DIR / "SI.pdf"
EXTERNAL_ROOT = BENCHMARK_ROOT / "external" / "northern_ghana"
RESULT_DIR = EXTERNAL_ROOT / "results"
FIGURE_DIR = BENCHMARK_ROOT / "figures"

ION_MOLAR_MASSES = {
    "Ca": 40.08,
    "Mg": 24.31,
    "Na": 22.99,
    "K": 39.10,
    "HCO3": 61.02,
    "Cl": 35.45,
    "SO4": 96.06,
    "NO3": 62.00,
    "F": 19.00,
    "Fe": 55.85,
    "PO4": 94.97,
}
ION_CHARGES = {
    "Ca": 2.0,
    "Mg": 2.0,
    "Na": 1.0,
    "K": 1.0,
    "HCO3": -1.0,
    "Cl": -1.0,
    "SO4": -2.0,
    "NO3": -1.0,
    "F": -1.0,
}
ION_ORDER = ["Ca", "Mg", "Na", "K", "HCO3", "Cl", "SO4", "NO3", "F", "Fe", "PO4"]
SOURCE_ION_ORDER = ["Ca", "Mg", "Na", "K", "HCO3", "Cl", "SO4", "NO3", "F"]
REQUIRED_COLUMNS = [
    "Well_ID",
    "Region",
    "District",
    "Community_Code",
    "Latitude",
    "Longitude",
    "Elevation_m",
    "Borehole_Depth_m",
    "Static_Water_Level_m",
    "Distance_River_km",
    "Distance_Farm_km",
    "Distance_Settlement_km",
    "pH",
    "EC_uS_cm",
    "TDS_mg_L",
    "Temperature_C",
    "Ca_mg_L",
    "Mg_mg_L",
    "Na_mg_L",
    "K_mg_L",
    "HCO3_mg_L",
    "Cl_mg_L",
    "SO4_mg_L",
    "NO3_mg_L",
    "F_mg_L",
    "d18O_permil",
    "d2H_permil",
    "Sr_mg_L",
    "SiO2_mg_L",
]
NESTED_RESULT_FIELDS = [
    "transport_probabilities",
    "candidate_scores",
    "constraints_active",
    "si_u",
    "si_v",
    "isotope_metrics",
    "gibbs_metrics",
    "qc_flags",
    "reaction_fit",
]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_safe(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return None if np.isnan(value) else float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    if pd.isna(value):
        return None
    return value


def _require_sources() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing corrected Northern Ghana workbook: {WORKBOOK_PATH}")


def _read_season_sheet(season: str) -> pd.DataFrame:
    frame = pd.read_excel(WORKBOOK_PATH, sheet_name=season)
    frame.columns = [str(column).strip() for column in frame.columns]
    missing = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"Sheet '{season}' is missing required columns: {missing}")

    frame = frame[REQUIRED_COLUMNS].copy()
    frame["Well_ID"] = frame["Well_ID"].astype(str).str.strip()
    frame["Season"] = season
    frame["Sample_ID"] = frame["Well_ID"] + "-" + season.lower()

    numeric_columns = [column for column in REQUIRED_COLUMNS if column != "Well_ID" and column not in {"Region", "District", "Community_Code"}]
    for column in numeric_columns:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")

    required_numeric_missing = int(frame[numeric_columns].isna().sum().sum())
    if required_numeric_missing:
        raise ValueError(f"Sheet '{season}' contains {required_numeric_missing} missing or non-numeric required values.")

    frame["Charge_Balance_Error_pct"] = _charge_balance_error(frame)
    classes = frame["Charge_Balance_Error_pct"].apply(_classify_cbe)
    frame["Data_Class"] = [item[0] for item in classes]
    frame["Flag"] = [item[1] for item in classes]
    frame["Hydrochemical_Facies"] = frame.apply(_hydrochemical_facies, axis=1)
    frame["Dominant_Process"] = "not supplied in corrected workbook"
    return frame


def _charge_balance_error(frame: pd.DataFrame) -> pd.Series:
    cations = sum(
        pd.to_numeric(frame[f"{ion}_mg_L"], errors="coerce") / ION_MOLAR_MASSES[ion] * abs(ION_CHARGES[ion])
        for ion in ["Ca", "Mg", "Na", "K"]
    )
    anions = sum(
        pd.to_numeric(frame[f"{ion}_mg_L"], errors="coerce") / ION_MOLAR_MASSES[ion] * abs(ION_CHARGES[ion])
        for ion in ["HCO3", "Cl", "SO4", "NO3", "F"]
    )
    denominator = cations + anions
    return 100.0 * (cations - anions) / denominator.replace(0.0, np.nan)


def _classify_cbe(cbe: float) -> tuple[str, str]:
    absolute = abs(float(cbe))
    if absolute <= 5.0:
        return "Quantitative inverse modelling", "CBE <= 5%; used for quantitative fitting"
    if absolute <= 10.0:
        return "Screening only - CBE between 5% and 10%", "5% < CBE <= 10%; excluded from quantitative fitting"
    return "Exploratory only - CBE > 10%", "CBE > 10%; excluded from quantitative fitting"


def _hydrochemical_facies(row: pd.Series) -> str:
    ca = float(row["Ca_mg_L"]) / ION_MOLAR_MASSES["Ca"] * 2.0
    mg = float(row["Mg_mg_L"]) / ION_MOLAR_MASSES["Mg"] * 2.0
    nak = float(row["Na_mg_L"]) / ION_MOLAR_MASSES["Na"] + float(row["K_mg_L"]) / ION_MOLAR_MASSES["K"]
    hco3 = float(row["HCO3_mg_L"]) / ION_MOLAR_MASSES["HCO3"]
    cl = float(row["Cl_mg_L"]) / ION_MOLAR_MASSES["Cl"]
    so4_no3_f = (
        float(row["SO4_mg_L"]) / ION_MOLAR_MASSES["SO4"] * 2.0
        + float(row["NO3_mg_L"]) / ION_MOLAR_MASSES["NO3"]
        + float(row["F_mg_L"]) / ION_MOLAR_MASSES["F"]
    )
    cation = max({"Ca": ca, "Mg": mg, "NaK": nak}.items(), key=lambda item: item[1])[0]
    anion = max({"HCO3": hco3, "Cl": cl, "SO4NO3F": so4_no3_f}.items(), key=lambda item: item[1])[0]
    return f"{cation}-{anion}"


def load_workbook_tables() -> dict[str, pd.DataFrame]:
    _require_sources()
    dry = _read_season_sheet("Dry")
    wet = _read_season_sheet("Wet")
    chemistry = pd.concat([wet, dry], ignore_index=True)
    wells = dry[
        [
            "Well_ID",
            "Region",
            "District",
            "Community_Code",
            "Latitude",
            "Longitude",
            "Elevation_m",
            "Borehole_Depth_m",
            "Static_Water_Level_m",
            "Distance_River_km",
            "Distance_Farm_km",
            "Distance_Settlement_km",
        ]
    ].drop_duplicates("Well_ID")
    return {"wells": wells, "chemistry": chemistry, "dry": dry, "wet": wet}


def _season_metadata_consistent(dry: pd.DataFrame, wet: pd.DataFrame) -> bool:
    compare_columns = ["Latitude", "Longitude", "Elevation_m", "Borehole_Depth_m", "Static_Water_Level_m"]
    merged = dry[["Well_ID", *compare_columns]].merge(
        wet[["Well_ID", *compare_columns]],
        on="Well_ID",
        suffixes=("_dry", "_wet"),
    )
    if len(merged) != len(dry) or len(merged) != len(wet):
        return False
    for column in compare_columns:
        diff = (merged[f"{column}_dry"] - merged[f"{column}_wet"]).abs()
        if bool((diff > 1e-9).any()):
            return False
    return True


def profile_workbook(tables: dict[str, pd.DataFrame]) -> dict[str, object]:
    wells = tables["wells"]
    chemistry = tables["chemistry"]
    cbe = chemistry["Charge_Balance_Error_pct"].abs()
    wet_dry_counts = chemistry.groupby(["Well_ID", "Season"]).size().unstack(fill_value=0)

    return {
        "source_workbook_name": WORKBOOK_PATH.name,
        "n_wells": int(len(wells)),
        "n_hydrochemical_samples": int(len(chemistry)),
        "n_feature_rows": 0,
        "n_source_graph_edges": 0,
        "n_graph_edges": 0,
        "n_climate_records": 0,
        "n_regions": int(wells["Region"].nunique()),
        "wet_dry_per_well_complete": bool(
            ((wet_dry_counts.get("Wet", 0) == 1) & (wet_dry_counts.get("Dry", 0) == 1)).all()
        ),
        "dry_wet_metadata_consistent": _season_metadata_consistent(tables["dry"], tables["wet"]),
        "quantitative_samples": int((cbe <= 5.0).sum()),
        "screening_samples": int(((cbe > 5.0) & (cbe <= 10.0)).sum()),
        "exploratory_samples": int((cbe > 10.0).sum()),
        "median_abs_cbe_pct": float(cbe.median()),
        "max_abs_cbe_pct": float(cbe.max()),
        "sampling_date_min": "not supplied",
        "sampling_date_max": "not supplied",
        "source_sheets": "Dry; Wet",
    }


def build_samples(chemistry: pd.DataFrame) -> pd.DataFrame:
    head_proxy = chemistry["Elevation_m"] - chemistry["Static_Water_Level_m"]
    samples = pd.DataFrame(
        {
            "site_id": chemistry["Well_ID"],
            "sample_id": chemistry["Sample_ID"],
            "well_id": chemistry["Well_ID"],
            "region": chemistry["Region"],
            "district": chemistry["District"],
            "community_code": chemistry["Community_Code"],
            "season": chemistry["Season"],
            "sample_date": None,
            "aquifer": "not supplied in corrected workbook",
            "latitude": chemistry["Latitude"],
            "longitude": chemistry["Longitude"],
            "lat": chemistry["Latitude"],
            "lon": chemistry["Longitude"],
            "elevation": chemistry["Elevation_m"],
            "well_depth_m": chemistry["Borehole_Depth_m"],
            "well_depth": chemistry["Borehole_Depth_m"],
            "dtw": chemistry["Static_Water_Level_m"],
            "head_proxy_m": head_proxy,
            "distance_river_km": chemistry["Distance_River_km"],
            "distance_farm_km": chemistry["Distance_Farm_km"],
            "distance_settlement_km": chemistry["Distance_Settlement_km"],
            "pH": chemistry["pH"],
            "EC": chemistry["EC_uS_cm"],
            "TDS": chemistry["TDS_mg_L"],
            "temp_c": chemistry["Temperature_C"],
            "d18O_permil": chemistry["d18O_permil"],
            "d2H_permil": chemistry["d2H_permil"],
            "Sr_mg_L": chemistry["Sr_mg_L"],
            "SiO2_mg_L": chemistry["SiO2_mg_L"],
            "calcite_si": np.nan,
            "dolomite_si": np.nan,
            "gypsum_si": np.nan,
            "halite_si": np.nan,
            "charge_balance_error_pct": chemistry["Charge_Balance_Error_pct"],
            "data_class": chemistry["Data_Class"],
            "dominant_process": chemistry["Dominant_Process"],
            "hydrochemical_facies": chemistry["Hydrochemical_Facies"],
            "recharge_zone_score": 1.0 / (1.0 + chemistry["Distance_River_km"]),
            "aridity_stress": np.nan,
            "graph_node_class": "corrected_workbook_well",
            "source_flag": chemistry["Flag"],
        }
    )
    for ion, molar_mass in ION_MOLAR_MASSES.items():
        source = f"{ion}_mg_L"
        if source in chemistry.columns:
            values = pd.to_numeric(chemistry[source], errors="coerce")
            samples[ion] = values / molar_mass
            samples[f"{ion}_mg_L"] = values
        else:
            samples[ion] = 0.0
            samples[f"{ion}_mg_L"] = 0.0
    return samples


def build_season_edges(
    season_samples: pd.DataFrame,
    season: str,
    max_edges: int,
    config,
) -> tuple[list[Edge], pd.DataFrame]:
    hydrosheaf_edges = infer_edges(
        season_samples.to_dict("records"),
        method="probabilistic",
        config=config,
    )
    rows = []
    edges = []
    for edge in hydrosheaf_edges:
        attrs = dict(edge.attrs or {})
        source_edge_id = edge.edge_id
        edge.edge_id = f"{season}:{source_edge_id}"
        attrs["hydrosheaf_infer_edges_method"] = "probabilistic"
        attrs["source_edge_id"] = source_edge_id
        edge.attrs = attrs
        edges.append(edge)
        edge_confidence = attrs.get("edge_confidence", attrs.get("p_uv"))
        distance_km = attrs.get("distance_km")
        delta_h = attrs.get("delta_h")
        rows.append(
            {
                "season": season,
                "edge_id": edge.edge_id,
                "hydrosheaf_source_edge_id": source_edge_id,
                "u": edge.u,
                "v": edge.v,
                "edge_type": edge.type,
                "distance_km": float(distance_km) if distance_km is not None else np.nan,
                "delta_h_m": float(delta_h) if delta_h is not None else np.nan,
                "sigma_delta_h_m": attrs.get("sigma_delta_h"),
                "p_uv": float(attrs.get("p_uv")) if attrs.get("p_uv") is not None else np.nan,
                "edge_confidence": float(edge_confidence) if edge_confidence is not None else np.nan,
                "composite_graph_weight": float(edge_confidence) if edge_confidence is not None else np.nan,
                "source_tier": attrs.get("source_tier"),
                "flags": attrs.get("flags"),
                "shared_aquifer": "not supplied",
                "expected_process_link": "not supplied",
                "direction_basis": "Hydrosheaf probabilistic edge inference using head proxy, distance, gradient, and confidence threshold",
            }
        )

    edge_inputs = pd.DataFrame(rows)
    if edge_inputs.empty:
        raise ValueError(
            f"Hydrosheaf graph inference produced no usable edges for the {season} season. "
            "Check edge_radius_km, edge_p_min, and the coordinate/head-proxy fields."
        )
    edge_inputs = (
        edge_inputs.sort_values(["composite_graph_weight", "distance_km"], ascending=[False, True])
        .drop_duplicates(["season", "u", "v"])
        .reset_index(drop=True)
    )
    if max_edges > 0:
        edge_inputs = edge_inputs.head(max_edges).copy()
        keep = set(edge_inputs["edge_id"].astype(str))
        edges = [edge for edge in edges if edge.edge_id in keep]
    return edges, edge_inputs


def hydrosheaf_config():
    config = default_config()
    config.ion_order = ION_ORDER.copy()
    config.weights = [1.0] * len(config.ion_order)
    config.conservative_weights = [0.01] * len(config.ion_order)
    config.conservative_weights[config.ion_order.index("Cl")] = 1.0
    config.active_minerals = ["calcite", "dolomite", "gypsum", "halite"]
    config.latent_endmembers_enabled = False
    config.latent_endmembers_count = 0
    config.sheaf_age_enabled = False
    config.sheaf_isotope_enabled = False
    config.isotope_enabled = False
    config.phreeqc_enabled = False
    return config


def graph_inference_config(edge_radius_km: float, edge_p_min: float, edge_max_neighbors: int, edge_head_inference: str):
    config = hydrosheaf_config()
    config.edge_radius_km = float(edge_radius_km)
    config.edge_p_min = float(edge_p_min)
    config.edge_max_neighbors = int(edge_max_neighbors)
    config.edge_head_inference = edge_head_inference
    config.edge_head_key = "head_meas"
    config.edge_dtw_key = "dtw"
    config.edge_elevation_key = "elevation"
    config.edge_well_depth_key = "well_depth"
    config.edge_aquifer_key = "aquifer_unit"
    config.edge_gradient_min = 1e-4
    config.edge_sigma_meas = 0.5
    config.edge_sigma_dtw = 1.0
    config.edge_sigma_elev = 1.0
    config.edge_sigma_topo = 10.0
    config.validate()
    return config


def run_fits(
    samples: pd.DataFrame,
    max_edges_per_season: int,
    edge_radius_km: float,
    edge_p_min: float,
    edge_max_neighbors: int,
    edge_head_inference: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    result_rows = []
    edge_input_frames = []
    fit_config = hydrosheaf_config()
    edge_config = graph_inference_config(edge_radius_km, edge_p_min, edge_max_neighbors, edge_head_inference)

    for season in ["Wet", "Dry"]:
        season_samples = samples[
            (samples["season"] == season) & (samples["data_class"] == "Quantitative inverse modelling")
        ].copy()
        season_samples["site_id"] = season_samples["well_id"]
        season_samples["sample_id"] = season_samples["well_id"]
        edges, edge_inputs = build_season_edges(season_samples, season, max_edges_per_season, edge_config)
        edge_input_frames.append(edge_inputs)
        print(f"{season}: fitting {len(edges)} Hydrosheaf-inferred graph edges over {len(season_samples)} quantitative samples.")
        results, _extras = fit_network_pipeline(season_samples.to_dict("records"), edges, fit_config)
        for result in results:
            row = asdict(result)
            row["season"] = season
            for label, value in zip(result.z_labels, result.z_extents):
                row[f"reaction_{label}"] = value
            for nested in NESTED_RESULT_FIELDS:
                row.pop(nested, None)
            result_rows.append(row)

    return pd.DataFrame(result_rows), pd.concat(edge_input_frames, ignore_index=True)


def summarize_outputs(
    profile: dict[str, object],
    samples: pd.DataFrame,
    edge_inputs: pd.DataFrame,
    results: pd.DataFrame,
    edge_radius_km: float,
    edge_p_min: float,
    edge_max_neighbors: int,
    edge_head_inference: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if results.empty:
        raise ValueError("Northern Ghana validation produced no candidate edge fits.")
    top20 = results.nsmallest(20, "objective_score")
    summary = {
        **profile,
        "n_graph_edges": int(len(edge_inputs)),
        "n_generated_graph_edges": int(len(edge_inputs)),
        "generated_graph_method": "Hydrosheaf infer_edges(method='probabilistic')",
        "generated_graph_edge_radius_km": float(edge_radius_km),
        "generated_graph_edge_p_min": float(edge_p_min),
        "generated_graph_edge_max_neighbors": int(edge_max_neighbors),
        "generated_graph_edge_head_inference": edge_head_inference,
        "fit_samples_quantitative": int(samples[samples["data_class"] == "Quantitative inverse modelling"].shape[0]),
        "candidate_edge_fits": int(len(results)),
        "wet_candidate_edge_fits": int((results["season"] == "Wet").sum()),
        "dry_candidate_edge_fits": int((results["season"] == "Dry").sum()),
        "top20_min_chemistry_r2": float(top20["chemistry_r2"].min()),
        "median_chemistry_r2": float(results["chemistry_r2"].median()),
        "median_objective_score": float(results["objective_score"].median()),
        "median_graph_weight_used": float(edge_inputs["composite_graph_weight"].median()),
        "median_generated_edge_distance_km": float(edge_inputs["distance_km"].median()),
    }
    qc_rows = [
        {
            "check": "required_workbook_missing_values",
            "value": int(samples[["well_id", "season", *SOURCE_ION_ORDER]].isna().sum().sum()),
            "interpretation": "Required well, season, and converted major-ion fields are complete when value is 0.",
        },
        {
            "check": "quantitative_cbe_samples",
            "value": int(profile["quantitative_samples"]),
            "interpretation": "Samples with abs(CBE) <= 5% used for fitted validation.",
        },
        {
            "check": "screening_only_cbe_samples",
            "value": int(profile["screening_samples"]),
            "interpretation": "Kept in sample export, excluded from quantitative fitting.",
        },
        {
            "check": "exploratory_cbe_samples",
            "value": int(profile["exploratory_samples"]),
            "interpretation": "Kept in sample export, excluded from quantitative fitting.",
        },
        {
            "check": "source_graph_edges_supplied",
            "value": int(profile["n_source_graph_edges"]),
            "interpretation": "The corrected workbook has no supplied graph-edge sheet; graph edges are generated reproducibly.",
        },
        {
            "check": "generated_graph_edges_used",
            "value": int(len(edge_inputs)),
            "interpretation": "Seasonal graph edges inferred by Hydrosheaf infer_edges(method='probabilistic') and passed to sparse fitting.",
        },
        {
            "check": "dry_wet_metadata_consistent",
            "value": int(bool(profile["dry_wet_metadata_consistent"])),
            "interpretation": "Dry and wet rows have matching coordinates, elevations, depths, and static water levels.",
        },
    ]
    return pd.DataFrame([summary]), pd.DataFrame(qc_rows)


def write_report(summary: pd.Series, results: pd.DataFrame) -> None:
    top = results.nsmallest(20, "objective_score")
    summary_cols = ["season", "edge_id", "transport_model", "objective_score", "chemistry_r2"]
    lines = [
        "# E4c Northern Ghana Field-Hydrochemistry Validation Report",
        "",
        f"Run timestamp: {time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())}",
        "",
        "## Source",
        "",
        "- Workbook: `data/NorthenGhana/NorthernGhana.xlsx`.",
        "- Sheets used: `Dry` and `Wet`.",
        "- Supplementary information: `data/NorthenGhana/SI.pdf` if retained for provenance.",
        "- Public repository/DOI status: not embedded in the corrected workbook; cite the final public source in manuscript text if required.",
        "",
        "## Dataset Profile",
        "",
        f"- Wells: {int(summary['n_wells'])}.",
        f"- Hydrochemical records: {int(summary['n_hydrochemical_samples'])}.",
        f"- Quantitative CBE records used for fitting: {int(summary['fit_samples_quantitative'])}.",
        f"- Source graph edges supplied: {int(summary['n_source_graph_edges'])}.",
        f"- Generated graph edges used: {int(summary['n_generated_graph_edges'])}.",
        f"- Candidate edge fits run: {int(summary['candidate_edge_fits'])}.",
        f"- Top-20 minimum chemistry R2: {float(summary['top20_min_chemistry_r2']):.3f}.",
        f"- Median chemistry R2: {float(summary['median_chemistry_r2']):.3f}.",
        "",
        "## Top 20 Inferred Connectivity",
        "",
        top[summary_cols].to_markdown(index=False),
        "",
        "## Reviewer Interpretation",
        "",
        "This Northern Ghana run replaces the local `manu.xlsx` pilot for M2. It validates Hydrosheaf's ability to ingest the corrected field workbook, compute charge-balance screening, harmonize major ions to mmol/L, infer graph priors with Hydrosheaf's probabilistic graph mechanism, and fit sparse hydrochemical process edges under wet/dry seasonal conditions.",
        "",
        "Guardrail: this is field-hydrochemistry and data-limited workflow evidence. The corrected workbook does not contain an independent process-truth graph, tracer-age output, MODPATH pathline truth, or external PHREEQC benchmark solution, so it should not be used to replace E1, E2, or E3.",
        "",
    ]
    (RESULT_DIR / "e4c_northern_ghana_report.md").write_text("\n".join(lines), encoding="utf-8")


def plot_e4_network(results: pd.DataFrame) -> None:
    top = results.nsmallest(15, "objective_score").copy()
    graph = nx.DiGraph()
    for row in top.itertuples(index=False):
        graph.add_edge(row.u, row.v, weight=float(row.chemistry_r2), season=row.season)

    fig, axes = plt.subplots(1, 2, figsize=(12.4, 4.9), gridspec_kw={"width_ratios": [1.15, 1.0]})
    ax = axes[0]
    pos = nx.spring_layout(graph, seed=17, k=1.08, iterations=150)
    season_colors = {"Wet": "#2563eb", "Dry": "#f59e0b"}
    edge_colors = [season_colors.get(graph.edges[edge].get("season"), "#64748b") for edge in graph.edges]
    nx.draw_networkx_nodes(graph, pos, node_color="#dcfce7", edgecolors="#1f2937", node_size=580, ax=ax)
    nx.draw_networkx_edges(
        graph,
        pos,
        arrows=True,
        arrowstyle="-|>",
        arrowsize=12,
        edge_color=edge_colors,
        width=1.35,
        ax=ax,
    )
    nx.draw_networkx_labels(graph, pos, font_size=6.2, ax=ax)
    ax.set_title("Top Northern Ghana generated graph-edge fits")
    handles = [
        plt.Line2D([0], [0], color=color, lw=2.2, label=season)
        for season, color in season_colors.items()
    ]
    ax.legend(handles=handles, loc="lower left", frameon=False, fontsize=8)
    ax.axis("off")

    ax = axes[1]
    top10 = top.head(10).sort_values("chemistry_r2")
    labels = [f"{row.season}: {row.u}->{row.v}" for row in top10.itertuples(index=False)]
    ax.barh(labels, top10["chemistry_r2"], color="#0f766e")
    lower = max(0.0, min(0.90, float(top10["chemistry_r2"].min()) - 0.01))
    ax.set_xlim(lower, 1.0)
    ax.set_title("Chemistry fit for top edges")
    ax.set_xlabel("Chemistry R2")
    fig.tight_layout()
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(FIGURE_DIR / "figure_s8_e4_sparse_pilot_network.png", dpi=300, bbox_inches="tight")
    plt.close(fig)


def write_manifest(
    profile: dict[str, object],
    summary: pd.Series,
    max_edges_per_season: int,
    edge_radius_km: float,
    edge_p_min: float,
    edge_max_neighbors: int,
    edge_head_inference: str,
) -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    props = workbook.properties
    source_summary = {
        "workbook": str(WORKBOOK_PATH),
        "supplementary_information": str(SI_PATH) if SI_PATH.exists() else None,
        "workbook_sha256": file_sha256(WORKBOOK_PATH),
        "si_sha256": file_sha256(SI_PATH) if SI_PATH.exists() else None,
        "workbook_creator": props.creator,
        "workbook_created": str(props.created),
        "workbook_modified": str(props.modified),
        "source_url_or_doi": "not embedded in corrected workbook; cite final public source in manuscript text",
        "profile": profile,
    }
    workbook.close()
    manifest = {
        "run_timestamp_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "validation_tier": "E4c_northern_ghana_field_hydrochemistry",
        "source_summary": source_summary,
        "hydrosheaf_config": {
            "ion_order": ION_ORDER,
            "active_minerals": ["calcite", "dolomite", "gypsum", "halite"],
            "edge_direction_source": "Hydrosheaf probabilistic graph inference using elevation, static water level, coordinates, and head uncertainty",
            "graph_generation": "hydrosheaf.inference.network_fit.infer_edges(method='probabilistic')",
            "edge_radius_km": float(edge_radius_km),
            "edge_p_min": float(edge_p_min),
            "edge_max_neighbors": int(edge_max_neighbors),
            "edge_head_inference": edge_head_inference,
            "max_edges_per_season": int(max_edges_per_season),
            "cbe_filter": "Data_Class == 'Quantitative inverse modelling'",
        },
        "metrics": summary.to_dict(),
        "outputs": {
            "samples": str(RESULT_DIR / "northern_ghana_samples.csv"),
            "edge_inputs": str(RESULT_DIR / "northern_ghana_edge_inputs.csv"),
            "edge_results": str(RESULT_DIR / "northern_ghana_edge_results.csv"),
            "summary": str(RESULT_DIR / "northern_ghana_validation_summary.csv"),
            "qc": str(RESULT_DIR / "northern_ghana_qc_summary.csv"),
            "report": str(RESULT_DIR / "e4c_northern_ghana_report.md"),
            "figure": str(FIGURE_DIR / "figure_s8_e4_sparse_pilot_network.png"),
        },
    }
    (RESULT_DIR / "e4c_northern_ghana_source_manifest.json").write_text(
        json.dumps(_json_safe(manifest), indent=2),
        encoding="utf-8",
    )


def run_northern_ghana(
    max_edges_per_season: int,
    edge_radius_km: float,
    edge_p_min: float,
    edge_max_neighbors: int,
    edge_head_inference: str,
) -> dict[str, object]:
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    tables = load_workbook_tables()
    profile = profile_workbook(tables)
    samples = build_samples(tables["chemistry"])
    results, edge_inputs = run_fits(
        samples,
        max_edges_per_season,
        edge_radius_km,
        edge_p_min,
        edge_max_neighbors,
        edge_head_inference,
    )
    results = results.sort_values("objective_score").reset_index(drop=True)
    summary_df, qc_df = summarize_outputs(
        profile,
        samples,
        edge_inputs,
        results,
        edge_radius_km,
        edge_p_min,
        edge_max_neighbors,
        edge_head_inference,
    )

    samples.to_csv(RESULT_DIR / "northern_ghana_samples.csv", index=False)
    edge_inputs.to_csv(RESULT_DIR / "northern_ghana_edge_inputs.csv", index=False)
    results.to_csv(RESULT_DIR / "northern_ghana_edge_results.csv", index=False)
    summary_df.to_csv(RESULT_DIR / "northern_ghana_validation_summary.csv", index=False)
    qc_df.to_csv(RESULT_DIR / "northern_ghana_qc_summary.csv", index=False)
    write_report(summary_df.iloc[0], results)
    plot_e4_network(results)
    write_manifest(
        profile,
        summary_df.iloc[0],
        max_edges_per_season,
        edge_radius_km,
        edge_p_min,
        edge_max_neighbors,
        edge_head_inference,
    )
    return summary_df.iloc[0].to_dict()


def main() -> None:
    parser = argparse.ArgumentParser(description="Run E4c Northern Ghana field-hydrochemistry validation.")
    parser.add_argument(
        "--max-edges-per-season",
        type=int,
        default=0,
        help="Limit to the highest-weight generated graph edges per season. Use 0 for all valid generated edges.",
    )
    parser.add_argument(
        "--edge-radius-km",
        type=float,
        default=50.0,
        help="Hydrosheaf probabilistic graph search radius in kilometers.",
    )
    parser.add_argument(
        "--edge-p-min",
        type=float,
        default=0.60,
        help="Minimum Hydrosheaf probabilistic flow-direction confidence.",
    )
    parser.add_argument(
        "--edge-max-neighbors",
        type=int,
        default=3,
        help="Maximum Hydrosheaf primary graph neighbors per upstream node.",
    )
    parser.add_argument(
        "--edge-head-inference",
        choices=["heuristic", "bayesian", "bayesian_mcmc"],
        default="heuristic",
        help="Hydrosheaf head inference mode for graph construction.",
    )
    args = parser.parse_args()
    summary = run_northern_ghana(
        args.max_edges_per_season,
        args.edge_radius_km,
        args.edge_p_min,
        args.edge_max_neighbors,
        args.edge_head_inference,
    )
    print(
        "E4c Northern Ghana run complete: "
        f"{int(summary['candidate_edge_fits'])} candidate fits; "
        f"top-20 minimum chemistry R2={float(summary['top20_min_chemistry_r2']):.3f}."
    )


if __name__ == "__main__":
    main()
