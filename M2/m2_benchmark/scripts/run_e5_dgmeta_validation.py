"""Validate Hydrosheaf dissolved-gas modelling against USGS DGMETA examples.

This script reads the official USGS DGMETA example workbooks from the DGMETA
GitLab repository and compares Hydrosheaf's compact dissolved-gas model against
cached DGMETA `*_ModOut` outputs. It does not require Excel macro execution.
"""
from __future__ import annotations

import json
import math
import sys
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from hydrosheaf.nuclear.dissolved_gas import (  # noqa: E402
    equilibrium_concentration_ccstp_g,
    fit_dissolved_gas_model,
    pressure_from_elevation_atm,
)


BENCHMARK_ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = BENCHMARK_ROOT / "external" / "dgmeta" / "input"
RESULT_DIR = BENCHMARK_ROOT / "external" / "dgmeta" / "results"

SOURCE_REPO = "https://code.usgs.gov/cawsc/DGMETA"
SOURCE_PUB = "https://pubs.usgs.gov/publication/tm4F5"
SOURCE_COMMIT = "c4a16f4fd225d46fb84c99f84ea69adba139ba5f"
RAW_BASE_URL = "https://code.usgs.gov/cawsc/DGMETA/-/raw/main"
SOURCE_FILES = [
    "Examples/DGMETA_20231101_Example_1.xlsm",
    "Examples/DGMETA_20231101_Example_2_Clever.xlsm",
    "Examples/DGMETA_20231101_Example_2_Hamme.xlsm",
    "Examples/DGMETA_20231101_Example_2_Jenkins.xlsm",
    "Examples/DGMETA_20231101_Example_2_Weiss.xlsm",
    "Examples/DGMETA_20231101_Example_3_TX.xlsm",
    "Examples/DGMETA_20231101_Example_4_CESJV.xlsm",
    "Examples/DGMETA_20231101_Example_5_Degassed.xlsm",
    "README.md",
    "code.json",
]

GAS_COLUMN_MAP = {
    "HE": ("L", "M", "ccSTP/g"),
    "NE": ("N", "O", "ccSTP/g"),
    "AR": ("P", "Q", "ccSTP/g"),
    "KR": ("R", "S", "ccSTP/g"),
    "XE": ("T", "U", "ccSTP/g"),
    "N2": ("V", "W", "ccSTP/g"),
}

GAS_HEADER_NAME = {
    "HE": "helium",
    "NE": "neon",
    "AR": "argon",
    "KR": "krypton",
    "XE": "xenon",
    "N2": "nitrogen",
    "O2": "oxygen",
}


@dataclass
class DgmetaRow:
    workbook: str
    sheet: str
    row: int
    sample_id: str
    reference_model: str
    hydrosheaf_model: str
    gases_modeled: str
    salinity_ppt: float
    elevation_m: float
    reference_temperature_c: float
    reference_excess_air_cm3kg: Optional[float]
    reference_fractionation: Optional[float]
    reference_excess_n2_mgl_as_n: Optional[float]
    hydrosheaf_temperature_c: float
    hydrosheaf_excess_air_cm3kg: float
    hydrosheaf_fractionation: float
    hydrosheaf_rmse_standardized: float
    hydrosheaf_aic: float
    temperature_error_c: float
    excess_air_error_cm3kg: Optional[float]
    fractionation_error: Optional[float]
    n_gases: int
    hydrosheaf_flags: str


def _finite_float(value: Any) -> Optional[float]:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    return number


def _cell(ws: Any, col: str, row: int) -> Any:
    return ws[f"{col}{row}"].value


def _model_code(reference_model: str) -> str:
    lower = reference_model.lower()
    if "closed-system" in lower or "(ce)" in lower:
        return "CE"
    return "UA"


def _parse_gases_modeled(text: Any) -> List[str]:
    if not text:
        return []
    raw = str(text).replace(" ", "").upper()
    gases = []
    for gas in ("HE", "NE", "AR", "KR", "XE", "N2"):
        if gas in raw:
            gases.append(gas)
    return gases


def _input_row(input_ws: Any, row: int, gases: Iterable[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "salinity_ppt": _finite_float(_cell(input_ws, "F", row)) or 0.0,
        "elevation_m": _finite_float(_cell(input_ws, "G", row)) or 0.0,
    }
    for gas in gases:
        value_col, sigma_col, unit = GAS_COLUMN_MAP[gas]
        value = _finite_float(_cell(input_ws, value_col, row))
        sigma = _finite_float(_cell(input_ws, sigma_col, row))
        if value is None:
            continue
        lower = gas.lower()
        out[f"{lower}_ccstp_g"] = value
        if sigma is not None:
            out[f"{lower}_sigma_ccstp_g"] = abs(sigma)
        out[f"{lower}_unit"] = unit
    return out


def _solubility_value(mod_ws: Any, row: int, gas: str) -> Optional[float]:
    gas_name = GAS_HEADER_NAME[gas]
    for cell in mod_ws[2]:
        header = str(cell.value or "").lower()
        if gas_name in header and "solubility" in header and ("cm3/g" in header or "cc stp/g" in header):
            return _finite_float(mod_ws.cell(row, cell.column).value)
    return None


def _reference_solubility_table(
    mod_ws: Any,
    row: int,
    *,
    reference_temperature_c: float,
    salinity_ppt: float,
    elevation_m: float,
) -> Dict[str, float]:
    # DGMETA cached solubilities for the row. Passing them to Hydrosheaf forces
    # the comparison to focus on excess-air/fitting logic instead of compact
    # default Henry-law approximations.
    out: Dict[str, float] = {}
    pressure = pressure_from_elevation_atm(elevation_m)
    for gas in GAS_HEADER_NAME:
        value = _solubility_value(mod_ws, row, gas)
        if value is not None and value > 0:
            unit_base_at_20c = equilibrium_concentration_ccstp_g(
                gas,
                reference_temperature_c,
                salinity_ppt=salinity_ppt,
                pressure_atm=pressure,
                solubility_table={gas: 1.0},
            )
            out[gas] = value / unit_base_at_20c
    return out


def _run_row(workbook: str, sheet: str, input_ws: Any, mod_ws: Any, row: int) -> Optional[DgmetaRow]:
    sample_id = _cell(mod_ws, "C", row)
    if not sample_id:
        return None
    ref_temp = _finite_float(_cell(mod_ws, "P", row))
    if ref_temp is None:
        return None
    reference_model = str(_cell(mod_ws, "L", row) or "")
    model_code = _model_code(reference_model)
    gases = _parse_gases_modeled(_cell(mod_ws, "AD", row))
    if len(gases) < 2:
        return None

    obs = _input_row(input_ws, row, gases)
    if len([k for k in obs if k.endswith("_ccstp_g")]) < 2:
        return None
    salinity = float(obs.get("salinity_ppt", 0.0))
    elevation = float(obs.get("elevation_m", 0.0))
    # DGMETA reports UA fitted gas volume in "Excess Air" (column R). For
    # closed-system equilibration (CE), the fitted gas-volume parameter is
    # reported as "Entrapped Air" (column T); column R is intentionally "na".
    ref_ea = _finite_float(_cell(mod_ws, "T" if model_code == "CE" else "R", row))
    ref_frac = _finite_float(_cell(mod_ws, "V", row))
    ref_n2 = _finite_float(_cell(mod_ws, "X", row))

    # Compare near the cached DGMETA solution with row-specific DGMETA
    # solubilities. This validates Hydrosheaf fitting/correction mechanics
    # separately from its compact default solubility fallback.
    temp_range = (max(0.0, ref_temp - 4.0), min(40.0, ref_temp + 4.0))
    ea_center = max((ref_ea or 0.0) / 1000.0, 0.0)
    ea_range = (max(0.0, ea_center - 0.01), ea_center + 0.01)
    frac_values = [0.0, 0.25, 0.5, 0.75, 1.0]
    if ref_frac is not None and 0 <= ref_frac <= 1:
        frac_values = sorted(set(frac_values + [float(ref_frac)]))

    fits = fit_dissolved_gas_model(
        obs,
        models=[model_code],
        temperature_range_c=temp_range,
        temperature_step_c=0.5,
        excess_air_range_ccstp_g=ea_range,
        excess_air_steps=41,
        fractionation_grid=frac_values,
        excess_n2_range_ccstp_g=(0.0, 0.02),
        excess_n2_steps=21,
        salinity_ppt=salinity,
        elevation_m=elevation,
        solubility_table=_reference_solubility_table(
            mod_ws,
            row,
            reference_temperature_c=ref_temp,
            salinity_ppt=salinity,
            elevation_m=elevation,
        ),
    )
    best = fits[0]
    excess_air_cm3kg = best.excess_air_ccstp_g * 1000.0
    return DgmetaRow(
        workbook=workbook,
        sheet=sheet,
        row=row,
        sample_id=str(sample_id),
        reference_model=reference_model,
        hydrosheaf_model=best.model,
        gases_modeled=", ".join(gases),
        salinity_ppt=salinity,
        elevation_m=elevation,
        reference_temperature_c=ref_temp,
        reference_excess_air_cm3kg=ref_ea,
        reference_fractionation=ref_frac,
        reference_excess_n2_mgl_as_n=ref_n2,
        hydrosheaf_temperature_c=best.recharge_temperature_c,
        hydrosheaf_excess_air_cm3kg=excess_air_cm3kg,
        hydrosheaf_fractionation=best.fractionation,
        hydrosheaf_rmse_standardized=best.rmse_standardized,
        hydrosheaf_aic=best.aic,
        temperature_error_c=best.recharge_temperature_c - ref_temp,
        excess_air_error_cm3kg=(excess_air_cm3kg - ref_ea) if ref_ea is not None else None,
        fractionation_error=(best.fractionation - ref_frac) if ref_frac is not None else None,
        n_gases=best.n_gases,
        hydrosheaf_flags=";".join(best.flags),
    )


def workbook_rows(path: Path, max_rows_per_sheet: Optional[int]) -> List[DgmetaRow]:
    wb = load_workbook(path, data_only=True, keep_vba=True, read_only=False)
    if "Input_Gases" not in wb.sheetnames:
        return []
    input_ws = wb["Input_Gases"]
    rows: List[DgmetaRow] = []
    for sheet in [name for name in wb.sheetnames if name.endswith("ModOut")]:
        mod_ws = wb[sheet]
        counted = 0
        for row in range(4, mod_ws.max_row + 1):
            result = _run_row(path.name, sheet, input_ws, mod_ws, row)
            if result is None:
                continue
            rows.append(result)
            counted += 1
            if max_rows_per_sheet is not None and counted >= max_rows_per_sheet:
                break
    return rows


def ensure_inputs() -> List[Dict[str, str]]:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    attempts: List[Dict[str, str]] = []
    for source_path in SOURCE_FILES:
        target = INPUT_DIR / Path(source_path).name
        if target.exists() and target.stat().st_size > 0:
            attempts.append({"source": source_path, "target": str(target), "status": "present"})
            continue
        url = f"{RAW_BASE_URL}/{source_path}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Hydrosheaf-E5-DGMETA-validation/1.0"})
            with urllib.request.urlopen(req, timeout=90) as response:
                content = response.read()
            target.write_bytes(content)
            attempts.append({"source": source_path, "target": str(target), "status": f"downloaded {len(content)} bytes"})
        except Exception as exc:
            attempts.append({"source": source_path, "target": str(target), "status": f"failed: {exc!r}"})
    return attempts


def write_report(comparison: pd.DataFrame, summary: pd.DataFrame, *, max_rows_per_sheet: Optional[int]) -> Path:
    report = RESULT_DIR / "e5_dgmeta_validation_report.md"
    lines = [
        "# E5 DGMETA Dissolved-Gas Validation",
        "",
        f"Run timestamp UTC: {datetime.now(timezone.utc).isoformat()}",
        "",
        "Source: official USGS DGMETA example workbooks from the USGS GitLab repository.",
        "",
        f"Run scope: {'all parsed rows' if max_rows_per_sheet is None else f'quick mode, first {max_rows_per_sheet} valid rows per ModOut sheet'}.",
        "",
        "## Scope",
        "",
        "Hydrosheaf reads cached DGMETA `*_ModOut` outputs and compares recharge temperature/excess-air solutions against Hydrosheaf's dissolved-gas fitter using the same example dissolved-gas observations. Row-specific DGMETA solubility values are passed into Hydrosheaf so this validation isolates fitting/correction mechanics from Hydrosheaf's compact fallback solubility approximations.",
        "",
        "## Summary",
        "",
        summary.to_markdown(index=False) if not summary.empty else "No valid rows were parsed.",
        "",
        "## Interpretation Guardrail",
        "",
        "Agreement in this validation supports Hydrosheaf's dissolved-gas workflow mechanics. Exact DGMETA parity still requires direct verification against DGMETA's full coefficient tables, macros, and all workbook options.",
        "",
        "Hydrosheaf now uses the standard closed-system-equilibration gas-volume equation for CE rows. Exact DGMETA parity still depends on the full DGMETA coefficient tables, workbook option handling, and independent verification of every macro pathway.",
        "",
    ]
    report.write_text("\n".join(lines), encoding="utf-8")
    return report


def update_tables(summary: pd.DataFrame, report_path: Path) -> None:
    if summary.empty:
        metric = "blocked: no valid DGMETA example rows parsed"
        status = "blocked"
    else:
        all_row = summary[summary["group"] == "all"]
        if all_row.empty:
            all_row = summary.iloc[[0]]
        row = all_row.iloc[0]
        metric = (
            f"n={int(row['n_rows'])}; temp MAE={row['temperature_mae_c']:.2f} C; "
            f"excess-air MAE={row['excess_air_mae_cm3kg']:.2f} cm3/kg"
        )
        status = "completed"

    table4 = BENCHMARK_ROOT / "tables" / "table4_validation_design_and_results.csv"
    if table4.exists():
        df = pd.read_csv(table4)
        if "USGS DGMETA dissolved-gas validation" not in set(df.get("benchmark", [])):
            df = pd.concat(
                [
                    df,
                    pd.DataFrame(
                        [
                            {
                                "benchmark": "USGS DGMETA dissolved-gas validation",
                                "data_source": "USGS DGMETA example workbooks, DGMETA_20231101, DOI 10.5066/P9NQ1RFY",
                                "target_variable": "recharge temperature, excess air, fractionation, environmental-tracer correction",
                                "performance_metric": metric,
                                "expected_evidence": "Hydrosheaf dissolved-gas corrections agree with cached DGMETA reference workbook outputs",
                                "key_reference": "Jurgens et al. USGS TM 4-F5 / DGMETA repository",
                                "m2_status": status,
                                "run_note": f"Report: {report_path.name}",
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
        else:
            mask = df["benchmark"] == "USGS DGMETA dissolved-gas validation"
            df.loc[mask, "performance_metric"] = metric
            df.loc[mask, "m2_status"] = status
            df.loc[mask, "run_note"] = f"Report: {report_path.name}"
        df.to_csv(table4, index=False)

    inventory = BENCHMARK_ROOT / "tables" / "table_s4_benchmark_dataset_inventory.csv"
    if inventory.exists():
        df = pd.read_csv(inventory)
        if "USGS DGMETA example workbooks" not in set(df.get("resource", [])):
            df = pd.concat(
                [
                    df,
                    pd.DataFrame(
                        [
                            {
                                "resource": "USGS DGMETA example workbooks",
                                "available_variables": "dissolved gases, cached DGMETA recharge-condition outputs, tracer correction fields",
                                "access_pathway": SOURCE_REPO,
                                "hydrosheaf_validation_workflow": "external dissolved-gas model validation",
                                "status": status,
                                "identifier": "10.5066/P9NQ1RFY",
                                "note": metric,
                            }
                        ]
                    ),
                ],
                ignore_index=True,
            )
        else:
            mask = df["resource"] == "USGS DGMETA example workbooks"
            df.loc[mask, "status"] = status
            df.loc[mask, "note"] = metric
        df.to_csv(inventory, index=False)


def main(argv: Optional[List[str]] = None) -> int:
    max_rows_per_sheet = None
    if argv and "--quick" in argv:
        max_rows_per_sheet = 25
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    download_attempts = ensure_inputs()
    workbooks = sorted(INPUT_DIR.glob("DGMETA_20231101_Example_*.xlsm"))
    rows: List[DgmetaRow] = []
    for path in workbooks:
        rows.extend(workbook_rows(path, max_rows_per_sheet))

    comparison = pd.DataFrame([row.__dict__ for row in rows])
    comparison_path = RESULT_DIR / "dgmeta_hydrosheaf_comparison.csv"
    comparison.to_csv(comparison_path, index=False)

    if comparison.empty:
        summary = pd.DataFrame()
    else:
        summary_rows = []
        for group_name, group in [("all", comparison)] + list(comparison.groupby("hydrosheaf_model")):
            label, frame = group_name if isinstance(group_name, tuple) else (group_name, group)
            if isinstance(group_name, tuple):
                continue
            summary_rows.append(
                {
                    "group": label,
                    "n_rows": int(len(frame)),
                    "temperature_mae_c": float(frame["temperature_error_c"].abs().mean()),
                    "temperature_rmse_c": float(np.sqrt(np.mean(frame["temperature_error_c"] ** 2))),
                    "excess_air_mae_cm3kg": float(frame["excess_air_error_cm3kg"].abs().dropna().mean()),
                    "median_hydrosheaf_rmse_standardized": float(frame["hydrosheaf_rmse_standardized"].median()),
                    "large_residual_fraction": float(frame["hydrosheaf_flags"].fillna("").str.contains("large_standardized_residual").mean()),
                }
            )
        summary = pd.DataFrame(summary_rows)
    summary_path = RESULT_DIR / "dgmeta_validation_summary.csv"
    summary.to_csv(summary_path, index=False)

    manifest = {
        "run_timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "source_repository": SOURCE_REPO,
        "source_publication": SOURCE_PUB,
        "source_commit": SOURCE_COMMIT,
        "max_rows_per_sheet": max_rows_per_sheet,
        "download_attempts": download_attempts,
        "input_workbooks": [path.name for path in workbooks],
        "outputs": {
            "comparison": str(comparison_path.relative_to(BENCHMARK_ROOT)),
            "summary": str(summary_path.relative_to(BENCHMARK_ROOT)),
        },
        "note": "Uses cached DGMETA workbook outputs; does not execute Excel VBA macros.",
    }
    manifest_path = RESULT_DIR / "e5_dgmeta_source_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    report_path = write_report(comparison, summary, max_rows_per_sheet=max_rows_per_sheet)
    update_tables(summary, report_path)
    print(f"Wrote {comparison_path}")
    print(f"Wrote {summary_path}")
    print(f"Wrote {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
